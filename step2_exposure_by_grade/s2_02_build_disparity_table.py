#!/usr/bin/env python3
"""Build a colleague-ready Excel workbook from the inter-grade disparity results.

Input : ALL_cities_inter-grade-disparity-block.csv  (produced by inter_grade_disparity.R)
Output: ALL_cities_inter-grade-disparity_TABLE.xlsx
        - "Read me"      : metric definitions, sign conventions, headline numbers, refs
        - "By city"      : one row per analysed city; tau_b & Cliff's delta per veg layer
        - "Full results" : the raw long table (every city x vegetation type)
"""
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CSV  = os.path.join(HERE, "ALL_cities_inter-grade-disparity-block.csv")
XLSX = os.path.join(HERE, "ALL_cities_inter-grade-disparity_TABLE.xlsx")
DATE = "2026-06-08"
VEGS = ["all", "tree", "shrub", "grass"]

df = pd.read_csv(CSV)

# ---- headline numbers (vegetation = all) ------------------------------------
oa = df[(df.status == "ok") & (df.veg_type == "all")]
H = dict(
    n_cities      = df.city.nunique(),
    n_analysed    = oa.city.nunique(),
    n_blocks      = int(df[df.veg_type == "all"].n_total.sum()),
    n_disp        = int((oa.kendall_tau_b < 0).sum()),
    n_disp_sig    = int(((oa.kendall_tau_b < 0) & (oa.kendall_p_fdr < 0.05)).sum()),
    n_rev         = int((oa.kendall_tau_b > 0).sum()),
    n_rev_sig     = int(((oa.kendall_tau_b > 0) & (oa.kendall_p_fdr < 0.05)).sum()),
    med_tau       = round(oa.kendall_tau_b.median(), 3),
    med_cliff     = round(oa.cliff_delta_AvsD.median(), 3),
)

# ---- wide "by city" table (analysed cities only) ----------------------------
base = (df[df.veg_type == "all"]
        [["city", "status", "n_total", "n_A", "n_B", "n_C", "n_D",
          "kendall_tau_b", "kendall_p_fdr", "cliff_delta_AvsD", "AvsD_p_fdr"]]
        .rename(columns={"kendall_tau_b": "tau_all", "kendall_p_fdr": "q_tau_all",
                         "cliff_delta_AvsD": "cliff_all", "AvsD_p_fdr": "q_cliff_all"}))
wide = base[base.status == "ok"].drop(columns="status").copy()
for vt in ["tree", "shrub", "grass"]:
    s = (df[df.veg_type == vt][["city", "kendall_tau_b", "cliff_delta_AvsD"]]
         .rename(columns={"kendall_tau_b": f"tau_{vt}", "cliff_delta_AvsD": f"cliff_{vt}"}))
    wide = wide.merge(s, on="city", how="left")
wide = wide.sort_values("tau_all").reset_index(drop=True)  # most disparity first
wide_cols = ["city", "n_total", "n_A", "n_B", "n_C", "n_D",
             "tau_all", "q_tau_all", "cliff_all", "q_cliff_all",
             "tau_tree", "cliff_tree", "tau_shrub", "cliff_shrub",
             "tau_grass", "cliff_grass"]
wide = wide[wide_cols]

full = df.sort_values(["veg_type", "kendall_tau_b"]).reset_index(drop=True)

# ---- styling helpers --------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 28

def numfmt(ws, header_row, name_to_fmt, nrow):
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for name, fmt in name_to_fmt.items():
        if name in headers:
            col = get_column_letter(headers[name])
            for r in range(2, nrow + 2):
                ws[f"{col}{r}"].number_format = fmt

# red = MORE disparity in both metrics (tau negative / cliff positive)
def scale_tau(ws, col, nrow):
    ws.conditional_formatting.add(f"{col}2:{col}{nrow+1}", ColorScaleRule(
        start_type="num", start_value=-0.4, start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.4, end_color="63BE7B"))
def scale_cliff(ws, col, nrow):
    ws.conditional_formatting.add(f"{col}2:{col}{nrow+1}", ColorScaleRule(
        start_type="num", start_value=-0.8, start_color="63BE7B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.8, end_color="F8696B"))

# ---- write workbook ---------------------------------------------------------
with pd.ExcelWriter(XLSX, engine="openpyxl") as xw:
    wide.to_excel(xw, sheet_name="By city", index=False)
    full.to_excel(xw, sheet_name="Full results", index=False)
    wb = xw.book
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---- Read me sheet (front) ----
    ws = wb.create_sheet("Read me", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 104
    title_font = Font(bold=True, size=14, color="1F4E78")
    sec_font   = Font(bold=True, size=11, color="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")
    rows = [
        ("title", "Greenspace disparity across HOLC redlining grades (block level)", ""),
        ("plain", "Generated", DATE),
        ("plain", "Source data", "ALL_cities_BLOCK-level-GE-redline.csv"),
        ("plain", "Unit of analysis", "U.S. Census block; greenspace = GE_all_* exposure"),
        ("plain", "Coverage",
         f"{H['n_blocks']:,} blocks (grades A-D) across {H['n_cities']} cities; "
         f"{H['n_analysed']} cities had ≥5 blocks in EVERY grade and were analysed "
         f"(the rest are flagged 'insufficient_n')."),
        ("sec", "Vegetation layers", ""),
        ("plain", "all", "GE_all_tot  (total vegetation)"),
        ("plain", "tree", "GE_all_Tr"),
        ("plain", "shrub", "GE_all_Sh"),
        ("plain", "grass", "GE_all_Gr"),
        ("sec", "Metric 1 - Kendall's tau-b", ""),
        ("plain", "What", "Rank correlation between HOLC grade (A=1, B=2, C=3, D=4) and "
                          "block greenspace. Range [-1, 1], scale-free, uses the full A<B<C<D order."),
        ("plain", "Sign", "NEGATIVE = disparity: greenspace DECLINES from grade A to grade D "
                          "(the redlining penalty). The more negative, the more systematic the gradient."),
        ("plain", "Significance", "kendall q = Benjamini-Hochberg FDR (across cities, within each "
                                  "vegetation layer). q < 0.05 = significant."),
        ("sec", "Metric 2 - Cliff's delta (A vs D)", ""),
        ("plain", "What", "Probability that a grade-A block is greener than a grade-D block, minus the "
                          "reverse. Range [-1, 1], scale-free. +1 = every A block greener than every D block."),
        ("plain", "Sign", "POSITIVE = disparity (grade A greener than grade D). NOTE: this is the "
                          "OPPOSITE sign to tau-b for the same gap (tau-b runs along A->D; Cliff is A vs D)."),
        ("plain", "Significance", "cliff q (AvsD q) = FDR-corrected Mann-Whitney test of grade A vs D."),
        ("sec", "Headline result (vegetation = all)", ""),
        ("plain", "Disparity direction",
         f"{H['n_disp']} of {H['n_analysed']} cities ({100*H['n_disp']/H['n_analysed']:.0f}%) have tau_b < 0 "
         f"(less greenspace in worse-graded areas); {H['n_disp_sig']} significant at q<0.05."),
        ("plain", "Reverse direction",
         f"{H['n_rev']} cities have tau_b > 0 ({H['n_rev_sig']} significant)."),
        ("plain", "Typical magnitude", f"median tau_b = {H['med_tau']}, median Cliff's delta = {H['med_cliff']}."),
        ("sec", "Reading the 'By city' sheet", ""),
        ("plain", "Sorting", "Rows sorted by tau_all ascending: most disparate cities at the top."),
        ("plain", "Colour", "Red = more disparity (tau negative / Cliff positive); green = reverse."),
        ("sec", "Key references", ""),
        ("plain", "Nardone et al. 2021", "Environ. Health Perspect. 129. doi:10.1289/EHP7495"),
        ("plain", "Locke et al. 2021", "npj Urban Sustainability 1. doi:10.1038/s42949-021-00022-0"),
        ("plain", "Schell et al. 2020", "Science 369. doi:10.1126/science.aay4497"),
        ("plain", "Hsu et al. 2021", "Nat. Commun. 12. doi:10.1038/s41467-021-22799-5"),
    ]
    r = 1
    for kind, a, b in rows:
        ca, cb = ws.cell(row=r, column=1, value=a), ws.cell(row=r, column=2, value=b)
        if kind == "title":
            ca.font = title_font; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        elif kind == "sec":
            ca.font = sec_font
        else:
            ca.font = Font(bold=True)
        cb.alignment = wrap
        r += 1

    # ---- style data sheets ----
    ws1 = xw.sheets["By city"]; n1 = len(wide)
    style_header(ws1, ws1.max_column)
    numfmt(ws1, 1, {**{f"tau_{v}": "0.000" for v in VEGS},
                    **{f"cliff_{v}": "0.000" for v in VEGS},
                    "q_tau_all": "0.0E+00", "q_cliff_all": "0.0E+00",
                    "n_total": "#,##0", "n_A": "#,##0", "n_B": "#,##0",
                    "n_C": "#,##0", "n_D": "#,##0"}, n1)
    hdr1 = {ws1.cell(row=1, column=c).value: get_column_letter(c) for c in range(1, ws1.max_column + 1)}
    for v in VEGS:
        scale_tau(ws1, hdr1[f"tau_{v}"], n1)
        scale_cliff(ws1, hdr1[f"cliff_{v}"], n1)
    ws1.column_dimensions["A"].width = 20
    for c in range(2, ws1.max_column + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 11

    ws2 = xw.sheets["Full results"]; n2 = len(full)
    style_header(ws2, ws2.max_column)
    numfmt(ws2, 1, {"kendall_tau_b": "0.000", "cliff_delta_AvsD": "0.000",
                    "kendall_p": "0.0E+00", "kendall_p_fdr": "0.0E+00",
                    "AvsD_p": "0.0E+00", "AvsD_p_fdr": "0.0E+00",
                    "n_total": "#,##0", "n_A": "#,##0", "n_B": "#,##0",
                    "n_C": "#,##0", "n_D": "#,##0"}, n2)
    ws2.column_dimensions["A"].width = 20

print(f"Wrote {XLSX}")
print(f"  By city: {len(wide)} cities | Full results: {len(full)} rows")
